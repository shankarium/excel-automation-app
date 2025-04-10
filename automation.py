import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def process_excel_fixed(input_path):
    input_df = pd.read_excel(input_path)
    reference_df = pd.read_excel("reference.xlsx")

    processed_rows = []
    for _, input_row in input_df.iterrows():
        subcat = input_row['Subcategory']
        sole = input_row['Soletype']
        design = input_row.get('Design', None)
        balance = input_row['Balance']

        match = reference_df[
            (reference_df['Subcategory'] == subcat) &
            (reference_df['Soletype'] == sole)
        ]
        if design:
            match = match[match['Design'] == design]

        if not match.empty:
            ref_row = match.iloc[0]
            try:
                sizes = [s.strip() for s in str(ref_row['Order']).split(',')]
                ratios = [float(r) for r in str(ref_row['Ration']).split(',')]
                total_ratio = sum(ratios)
                for size, ratio in zip(sizes, ratios):
                    qty = round(balance * ratio / total_ratio)
                    processed_rows.append({
                        'Article': input_row['Article'],
                        'Gender': input_row['Gender'],
                        'Colour': input_row['Colour'],
                        'Subcategory': subcat,
                        'Soletype': sole,
                        'Design': ref_row['Design'],
                        'Size': size,
                        'Qty': qty
                    })
            except Exception as e:
                print(f"Skipping row due to error: {e}")
        else:
            print(f"No match found for: {subcat}, {sole}, {design}")

    output_df = pd.DataFrame(processed_rows)
    output_path = "static/processed.xlsx"
    output_df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    red_font = Font(color="FF0000")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = red_font
    wb.save(output_path)

    return output_path
